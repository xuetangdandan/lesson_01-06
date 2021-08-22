#!/usr/bin/env python 
# -*- coding: utf-8 -*- 
# @Time : 2020/9/2 21:42 
# @Author : Lemon_Tricy
# @QQ: 2378807189
# Copyright：湖南省零檬信息技术有限公司
"""
接口测试自动化的步骤：
1、接口测试用例 --- Done
2、Python代码读取接口测试用例 --- Done--read_date()
3、requests 库发送接口请求 --- Done=== api_request
4、执行结果 vs  预期结果  == 用例执行是否是通过的！--结果
5、结果回写到excel里  -- openpyxl --Done -- write_result()
函数定义：
1、实现功能  2、 参数 --变化的值  3、 返回值--- 别人需要从你这里得到的数据

代码自动读取测试数据 + 自动化回写数据  === 测试用例一般excel里居多 ==操作excel
第三方库：openpyxl  --- 读取、回写
1、安装： pip install openpyxl
2、导入
excel表格的常用操作：三大对象
1、工作簿对象
2、表单--sheet
3、单元格 --cell
"""
import openpyxl  #导入第三方库 openpyxl
import requests
# 读取数据函数
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)   # 加载工作对象
    sheet = wb[sheetname]   # 获取到表单
    case_list = []  # 装用例的大列表
    max_row = sheet.max_row # 获取表单里的最大行数
    for i in range(2,max_row+1):
        case = dict(
        case_id = sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value,   # 行、列 找到对应的单元格
        data = sheet.cell(row=i,column=6).value,  # 参数
        expected = sheet.cell(row=i,column=7).value
        )  # 一个字典是一个测试用例
        case_list.append(case)
    return case_list  # 返回值

# 发送接口请求
def api_request(api_url,api_data):
    qcd_headers = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    response = requests.post(url=api_url,json=api_data,headers=qcd_headers)  # 返回值 --响应消息
    return response.json()  # 得到响应结果

# 回写数据
def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)   # 加载工作对象
    sheet = wb[sheetname]   # 获取到表单
    sheet.cell(row=row,column=column).value = final_result   # 数据重写
    wb.save(filename)  # 保存--- 写入的数据才会生效

"""
[
{
'case_id': 1, 
'url': 'http://api.lemonban.com/futureloan/member/register', 
'data': '{"mobile_phone":"13552440101","pwd":"12345678","type":1,"reg_name":"34254sdfs"}', 
'expected': '{"code": 0, "msg": "OK"}'
}, 
]
{'code': 2, 'msg': '账号已存在', 'data': None, 'copyright': 'Copyright 柠檬班 © 2017-2020 湖南省零檬信息技术有限公司 All Rights Reserved'}
"""
def execute_func(filename,sheetname):
    cases = read_data(filename,sheetname)  # 变量接受函数的返回值
    for case in cases:
        case_id = case.get("case_id")
        url = case.get("url")
        data = case["data"]
        # print(type(data))  # 字符串---- 转换为字典  -- -eval():运行字符串包裹的Python表达式 -- 脱掉引号的外衣
        data = eval(data)   # 脱引号
        # print(type(data))
        expected = case.get("expected")  # 字符串  --字典
        expected = eval(expected)
        real_result = api_request(api_url=url,api_data=data)   # 执行结果
        real_msg = real_result["msg"]
        expected_msg = expected.get("msg")
        print("执行结果是：{}".format(real_msg))
        print("预期结果是：{}".format(expected_msg))
        if real_msg == expected_msg:
            print("第{}条测试用例执行通过！".format(case_id))
            final_result = "Passed"
        else:
            print("第{}条测试用例执行不通过！".format(case_id))
            final_result = "Failed"
        print("*" * 20)
        write_result(filename,sheetname,case_id+1,8,final_result)
execute_func("test_case_api.xlsx","register")



















